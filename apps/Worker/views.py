from django.contrib import messages
from django.shortcuts import redirect, render, get_object_or_404
from django.shortcuts import render
from .models import Worker
from .forms import WorkerForm
from apps.WorkerDocuments.models import WorkerDocuments
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_workbook
from django.http import HttpResponse
import io

#Listar company de trabajos
@login_required
def list_worker(request):
    listado = Worker.objects.all();
    
    return render(request, 'Worker/list_worker.html', {'listado':listado})

#Formulario centro de trabajo
@login_required
def form_worker(request):

    data = {
        'form': WorkerForm(),
    }
    
    
    if request.method == 'POST':
        formulario_worker = WorkerForm (data=request.POST)
        if formulario_worker.is_valid():
            formulario_worker.save()
            messages.success(request, "Registro agregado correctamente")
            return redirect(to="list_worker")
        else:
            data["form"] = formulario_worker
            messages.error(request, "Error al guardar el formulario")
            

    return render(request, 'Worker/form_worker.html', data)

#Editar Trabajador
@login_required
def edit_worker(request,id):
    worker = get_object_or_404(Worker, id_worker=id)
    workerdocuments = WorkerDocuments.objects.filter(id_worker=id)
    if request.method == 'POST':
        form = WorkerForm (request.POST, instance=worker)
        if form.is_valid():
            form.save()
            messages.success(request, "Registro agregado correctamente")
            return redirect(to="list_worker")
        else:
            messages.error(request, "Error al editar el formulario")
    else:
        form = WorkerForm(instance=worker)
        
    context = {
        'form': form,
        'worker': worker,
        'workerdocuments': workerdocuments
    }
    return render(request, 'Worker/edit_worker.html', context)

#Editar eliminar Trabajador
@login_required
def delet_worker(request, id):
    worker = get_object_or_404(Worker, id_worker=id)
    worker.delete()
    messages.success(request, "Eliminado correctamente")
    return redirect(to="list_worker")

#Filtrado de Documentos del trabajador
@login_required
def worker_documents(request, worker_id):
    worker = get_object_or_404(Worker, id_worker=worker_id)
    workerdocuments = WorkerDocuments.objects.filter(id_worker_id=worker_id)
    context = {
        'worker': worker,
        'workerdocuments': workerdocuments,
    }
    return render(request, 'Worker/worker_documents.html', context)

def buscar_worker(request):
    query = request.GET.get('q')
    workers = Worker.objects.filter(name__icontains=query) if query else Worker.objects.all()
    
    context = {
        'listado': workers
    }
    
    return render(request, 'Worker/list_worker.html', context)



def listado_trabajadores(request):
    # Obtener el listado de trabajadores desde la base de datos
    listado_trabajadores = Worker.objects.all()

    if 'export' in request.GET:
        # Exportar a Excel si se recibió el parámetro 'export' en la solicitud GET
        workbook = Workbook()
        worksheet = workbook.active

        # Escribir encabezados de columna
        headers = ['Id', 'Nombres', 'Apellidos', 'RUT', 'Dirección', 'Región', 'Comuna', 'Teléfono', 'Email']
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            worksheet[f'{column_letter}1'] = header

        # Escribir datos de trabajadores
        row = 2
        for trabajador in listado_trabajadores:
            worksheet.cell(row=row, column=1, value=trabajador.id_worker)
            worksheet.cell(row=row, column=2, value=trabajador.name)
            worksheet.cell(row=row, column=3, value=trabajador.lastname)
            worksheet.cell(row=row, column=4, value=trabajador.rut)
            worksheet.cell(row=row, column=5, value=trabajador.address)
            worksheet.cell(row=row, column=6, value=str(trabajador.region))
            worksheet.cell(row=row, column=7, value=str(trabajador.comuna))
            worksheet.cell(row=row, column=8, value=trabajador.phone)
            worksheet.cell(row=row, column=9, value=trabajador.email)
            row += 1

        # Guardar el libro de Excel en un archivo virtual
        virtual_workbook = io.BytesIO()
        save_workbook(workbook, virtual_workbook)
        virtual_workbook.seek(0)

        # Crear la respuesta de descarga
        response = HttpResponse(virtual_workbook.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=trabajadores.xlsx'
        return response

    # Si no se solicita la exportación, renderizar la plantilla normalmente
    context = {
        'listado': listado_trabajadores
    }
    return render(request, 'listado_trabajadores.html', context)


def download_excel(request):
    # Obtener el listado de trabajadores desde la base de datos
    listado_trabajadores = Worker.objects.all()

    # Crear el libro de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Escribir encabezados de columna
    headers = ['Id', 'Nombres', 'Apellidos', 'RUT', 'Dirección', 'Región', 'Comuna', 'Teléfono', 'Email']
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet[f'{column_letter}1'] = header

    # Escribir datos de trabajadores
    row = 2
    for trabajador in listado_trabajadores:
        worksheet.cell(row=row, column=1, value=trabajador.id_worker)
        worksheet.cell(row=row, column=2, value=trabajador.name)
        worksheet.cell(row=row, column=3, value=trabajador.lastname)
        worksheet.cell(row=row, column=4, value=trabajador.rut)
        worksheet.cell(row=row, column=5, value=trabajador.address)
        worksheet.cell(row=row, column=6, value=str(trabajador.region))
        worksheet.cell(row=row, column=7, value=str(trabajador.comuna))
        worksheet.cell(row=row, column=8, value=trabajador.phone)
        worksheet.cell(row=row, column=9, value=trabajador.email)
        row += 1

    # Guardar el libro de Excel en un archivo virtual
    virtual_workbook = io.BytesIO()
    save_workbook(workbook, virtual_workbook)
    virtual_workbook.seek(0)

    # Crear la respuesta de descarga
    response = HttpResponse(virtual_workbook.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_trabajadores.xlsx"'

    return response

